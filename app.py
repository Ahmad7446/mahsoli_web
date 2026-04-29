"""
محصولي ويب — نظام إدارة القطف والمبيعات
Flask Web Application
"""
import os, json, sqlite3, uuid, shutil, re
from collections import defaultdict
from datetime import datetime, date
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for
import io

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'mahsoli-secret-2025')

# ── مسارات الملفات ──
_BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
_APP_DIR     = os.path.join(_BASE_DIR, 'mahsoli_data')
DB_FILE      = os.path.join(_APP_DIR, 'mahsoli.db')
SETTINGS_FILE = os.path.join(_APP_DIR, 'settings.json')
BACKUP_DIR   = os.path.join(_APP_DIR, 'backups')
REPORTS_DIR  = os.path.join(_APP_DIR, 'reports')
os.makedirs(BACKUP_DIR, exist_ok=True)
os.makedirs(REPORTS_DIR, exist_ok=True)

ARABIC_WEEKDAYS = ['الإثنين','الثلاثاء','الأربعاء','الخميس','الجمعة','السبت','الأحد']

DEFAULT_PRODUCTS = [
    "بندورة","خيار","فليفلة","باذنجان","كوسا","خس",
    "بقدونس","نعناع","ملفوف","قرنبيط","بصل","ثوم",
    "فراولة","بطيخ","شمام","ذرة","فول","عدس",
    "تفاح","عنب","زيتون","برتقال","ليمون","أخرى"
]

_settings_cache = None

def load_settings(force=False):
    global _settings_cache
    if _settings_cache and not force:
        return _settings_cache
    default = {
        "farm_name": "مزرعتي",
        "sections": {
            "motez":  {"display": "معتز",  "type": "full",   "table": "motez"},
            "waleed": {"display": "وليد",  "type": "shared", "table": "waleed"}
        },
        "products": DEFAULT_PRODUCTS,
        "worker_ratio": 0.333
    }
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                s = json.load(f)
            for k, v in default.items():
                if k not in s:
                    s[k] = v
            for key, sec in s.get('sections', {}).items():
                if 'table' not in sec:
                    raw = sec.get('filename', key)
                    safe = re.sub(r'[^a-zA-Z0-9_]', '_', raw)
                    sec['table'] = safe if safe else key
            _settings_cache = s
            return s
        except Exception:
            pass
    _settings_cache = default
    with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(default, f, ensure_ascii=False, indent=4)
    return default

def save_settings(s):
    global _settings_cache
    with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(s, f, ensure_ascii=False, indent=4)
    _settings_cache = s

def get_db():
    conn = sqlite3.connect(DB_FILE, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn

def new_id():
    return str(uuid.uuid4())

def safe_float(val, default=0.0):
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(val)
    except:
        return default

def arabic_day(d):
    if isinstance(d, str):
        try:
            d = date.fromisoformat(d)
        except:
            return ''
    return ARABIC_WEEKDAYS[d.weekday()]

def _ensure_column(conn, table, column, col_type, default_sql=None):
    try:
        cols = conn.execute(f"PRAGMA table_info({table})").fetchall()
        exists = any(c['name'] == column for c in cols)
        if not exists:
            sql = f"ALTER TABLE {table} ADD COLUMN {column} {col_type}"
            if default_sql:
                sql += f" DEFAULT {default_sql}"
            conn.execute(sql)
    except:
        pass

def init_db():
    settings = load_settings()
    conn = get_db()
    c = conn.cursor()
    for key, cfg in settings['sections'].items():
        tbl = cfg['table']
        c.execute(f"""
            CREATE TABLE IF NOT EXISTS {tbl} (
                id TEXT PRIMARY KEY,
                day TEXT, date TEXT, product TEXT,
                qty REAL, unit TEXT, price REAL, total REAL,
                buyer TEXT, accounted TEXT DEFAULT 'باقي',
                notes TEXT, created_at TEXT, season_id TEXT
            )
        """)
        c.execute(f"CREATE INDEX IF NOT EXISTS idx_{tbl}_date ON {tbl}(date)")
        c.execute(f"CREATE INDEX IF NOT EXISTS idx_{tbl}_buyer ON {tbl}(buyer)")
        c.execute(f"CREATE INDEX IF NOT EXISTS idx_{tbl}_accounted ON {tbl}(accounted)")
    c.execute("""CREATE TABLE IF NOT EXISTS advances (
        id TEXT PRIMARY KEY, date TEXT, amount REAL,
        product TEXT, reason TEXT, notes TEXT, created_at TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS buyers (
        id TEXT PRIMARY KEY, name TEXT UNIQUE, phone TEXT,
        address TEXT, credit_limit REAL DEFAULT 0,
        category TEXT DEFAULT 'عادي', notes TEXT, created_at TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS payments (
        id TEXT PRIMARY KEY, buyer TEXT, amount REAL,
        date TEXT, notes TEXT, section TEXT, created_at TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS expenses (
        id TEXT PRIMARY KEY, date TEXT, category TEXT, amount REAL,
        notes TEXT, section TEXT, created_at TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS seasons (
        id TEXT PRIMARY KEY, name TEXT NOT NULL,
        section_key TEXT, product TEXT DEFAULT 'الكل',
        start_date TEXT NOT NULL, end_date TEXT NOT NULL,
        status TEXT DEFAULT 'مغلق', closed_at TEXT,
        notes TEXT, created_at TEXT)""")
    _ensure_column(conn, 'seasons', 'product', 'TEXT', "'الكل'")
    _ensure_column(conn, 'seasons', 'status', 'TEXT', "'مغلق'")
    for key, cfg in settings['sections'].items():
        _ensure_column(conn, cfg['table'], 'season_id', 'TEXT')
    conn.commit()
    conn.close()

# ══════════════════════════════════════════
# ROUTES — الصفحات
# ══════════════════════════════════════════

@app.route('/')
def index():
    return redirect(url_for('dashboard'))

@app.route('/dashboard')
def dashboard():
    settings = load_settings()
    return render_template('dashboard.html', settings=settings)

@app.route('/sales/<section_key>')
def sales(section_key):
    settings = load_settings()
    if section_key not in settings['sections']:
        return redirect(url_for('dashboard'))
    section = settings['sections'][section_key]
    return render_template('sales.html', settings=settings,
                           section_key=section_key, section=section)

@app.route('/buyers')
def buyers():
    settings = load_settings()
    return render_template('buyers.html', settings=settings)

@app.route('/reports')
def reports():
    settings = load_settings()
    return render_template('reports.html', settings=settings)

@app.route('/settings_page')
def settings_page():
    settings = load_settings()
    return render_template('settings.html', settings=settings)

# ══════════════════════════════════════════
# API — البيانات
# ══════════════════════════════════════════

@app.route('/api/dashboard')
def api_dashboard():
    settings = load_settings()
    conn = get_db()
    result = {
        'sections': {},
        'total_sales': 0, 'total_paid': 0, 'total_unpaid': 0,
        'today_sales': 0, 'week_sales': 0,
        'top_buyers': [], 'recent': []
    }
    today = date.today().isoformat()
    week_start = date.today().replace(day=date.today().day - date.today().weekday()).isoformat()

    all_rows = []
    for key, cfg in settings['sections'].items():
        tbl = cfg['table']
        rows = conn.execute(f"SELECT * FROM {tbl} ORDER BY date DESC, created_at DESC").fetchall()
        rows = [dict(r) for r in rows]
        total = sum(r['total'] or 0 for r in rows)
        paid = sum(r['total'] or 0 for r in rows if r['accounted'] == 'محاسب')
        unpaid = total - paid
        today_sum = sum(r['total'] or 0 for r in rows if r['date'] == today)
        result['sections'][key] = {
            'display': cfg['display'],
            'total': total, 'paid': paid, 'unpaid': unpaid,
            'count': len(rows), 'today': today_sum
        }
        result['total_sales'] += total
        result['total_paid'] += paid
        result['total_unpaid'] += unpaid
        result['today_sales'] += today_sum
        result['week_sales'] += sum(r['total'] or 0 for r in rows if r['date'] >= week_start)
        all_rows.extend(rows)

    # أحدث 10 سجلات
    all_rows.sort(key=lambda x: (x.get('date',''), x.get('created_at','')), reverse=True)
    result['recent'] = all_rows[:10]

    # أكبر المشترين ديناً
    buyer_debt = defaultdict(float)
    for key, cfg in settings['sections'].items():
        tbl = cfg['table']
        rows = conn.execute(f"SELECT buyer, SUM(total) as s FROM {tbl} WHERE accounted!='محاسب' AND buyer!='' GROUP BY buyer").fetchall()
        for r in rows:
            if r['buyer']:
                buyer_debt[r['buyer']] += r['s'] or 0
    result['top_buyers'] = sorted([{'name': k, 'debt': v} for k, v in buyer_debt.items()], key=lambda x: -x['debt'])[:5]

    # إحصائيات المحاصيل
    product_stats = defaultdict(float)
    for r in all_rows:
        if r.get('product'):
            product_stats[r['product']] += r['total'] or 0
    result['products'] = sorted([{'name': k, 'total': v} for k, v in product_stats.items()], key=lambda x: -x['total'])[:8]

    conn.close()
    return jsonify(result)

@app.route('/api/sales/<section_key>', methods=['GET'])
def api_get_sales(section_key):
    settings = load_settings()
    if section_key not in settings['sections']:
        return jsonify({'error': 'قسم غير موجود'}), 404
    cfg = settings['sections'][section_key]
    tbl = cfg['table']
    conn = get_db()

    # فلاتر
    buyer = request.args.get('buyer', '')
    product = request.args.get('product', '')
    accounted = request.args.get('accounted', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    limit = int(request.args.get('limit', 200))

    q = f"SELECT * FROM {tbl} WHERE 1=1"
    params = []
    if buyer:
        q += " AND buyer=?"; params.append(buyer)
    if product:
        q += " AND product=?"; params.append(product)
    if accounted:
        q += " AND accounted=?"; params.append(accounted)
    if date_from:
        q += " AND date>=?"; params.append(date_from)
    if date_to:
        q += " AND date<=?"; params.append(date_to)
    q += " ORDER BY date DESC, created_at DESC LIMIT ?"
    params.append(limit)

    rows = [dict(r) for r in conn.execute(q, params).fetchall()]
    total = sum(r['total'] or 0 for r in rows)
    paid = sum(r['total'] or 0 for r in rows if r['accounted'] == 'محاسب')
    conn.close()
    return jsonify({'rows': rows, 'total': total, 'paid': paid, 'unpaid': total - paid})

@app.route('/api/sales/<section_key>', methods=['POST'])
def api_add_sale(section_key):
    settings = load_settings()
    if section_key not in settings['sections']:
        return jsonify({'error': 'قسم غير موجود'}), 404
    cfg = settings['sections'][section_key]
    tbl = cfg['table']
    data = request.json

    date_str = data.get('date', date.today().isoformat())
    qty = safe_float(data.get('qty', 0))
    price = safe_float(data.get('price', 0)) if data.get('price') else None
    total = qty * (price or 0)

    row_id = new_id()
    conn = get_db()
    conn.execute(f"""
        INSERT INTO {tbl} (id,day,date,product,qty,unit,price,total,buyer,accounted,notes,created_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    """, (row_id, arabic_day(date_str), date_str,
          data.get('product',''), qty, data.get('unit','كغ'),
          price, total, data.get('buyer',''),
          data.get('accounted','باقي'), data.get('notes',''),
          datetime.now().isoformat()))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': row_id, 'total': total})

@app.route('/api/sales/<section_key>/<row_id>', methods=['PUT'])
def api_update_sale(section_key, row_id):
    settings = load_settings()
    if section_key not in settings['sections']:
        return jsonify({'error': 'قسم غير موجود'}), 404
    cfg = settings['sections'][section_key]
    tbl = cfg['table']
    data = request.json

    date_str = data.get('date', date.today().isoformat())
    qty = safe_float(data.get('qty', 0))
    price = safe_float(data.get('price', 0)) if data.get('price') else None
    total = qty * (price or 0)

    conn = get_db()
    conn.execute(f"""
        UPDATE {tbl} SET day=?,date=?,product=?,qty=?,unit=?,price=?,total=?,
        buyer=?,accounted=?,notes=? WHERE id=?
    """, (arabic_day(date_str), date_str, data.get('product',''), qty,
          data.get('unit','كغ'), price, total, data.get('buyer',''),
          data.get('accounted','باقي'), data.get('notes',''), row_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'total': total})

@app.route('/api/sales/<section_key>/<row_id>', methods=['DELETE'])
def api_delete_sale(section_key, row_id):
    settings = load_settings()
    if section_key not in settings['sections']:
        return jsonify({'error': 'قسم غير موجود'}), 404
    cfg = settings['sections'][section_key]
    tbl = cfg['table']
    conn = get_db()
    conn.execute(f"DELETE FROM {tbl} WHERE id=?", (row_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/buyers', methods=['GET'])
def api_get_buyers():
    conn = get_db()
    buyers = [dict(b) for b in conn.execute("SELECT * FROM buyers ORDER BY name").fetchall()]
    settings = load_settings()
    # احسب الديون
    for b in buyers:
        debt = 0
        for cfg in settings['sections'].values():
            tbl = cfg['table']
            r = conn.execute(f"SELECT SUM(total) FROM {tbl} WHERE buyer=? AND accounted!='محاسب'", (b['name'],)).fetchone()
            debt += r[0] or 0
        paid_sum = conn.execute("SELECT SUM(amount) FROM payments WHERE buyer=?", (b['name'],)).fetchone()[0] or 0
        b['debt'] = debt
        b['paid_sum'] = paid_sum
    conn.close()
    return jsonify(buyers)

@app.route('/api/buyers', methods=['POST'])
def api_add_buyer():
    data = request.json
    conn = get_db()
    bid = new_id()
    try:
        conn.execute("""INSERT INTO buyers (id,name,phone,address,category,notes,created_at)
            VALUES (?,?,?,?,?,?,?)""",
            (bid, data['name'], data.get('phone',''), data.get('address',''),
             data.get('category','عادي'), data.get('notes',''), datetime.now().isoformat()))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': 'اسم المشتري موجود مسبقاً'}), 400
    conn.close()
    return jsonify({'success': True, 'id': bid})

@app.route('/api/buyers/<name>', methods=['PUT'])
def api_update_buyer(name):
    data = request.json
    conn = get_db()
    conn.execute("""UPDATE buyers SET phone=?,address=?,category=?,notes=? WHERE name=?""",
        (data.get('phone',''), data.get('address',''),
         data.get('category','عادي'), data.get('notes',''), name))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/buyers/<name>/mark_paid', methods=['POST'])
def api_mark_paid(name):
    data = request.json
    section_key = data.get('section_key', '')
    settings = load_settings()
    conn = get_db()
    if section_key and section_key in settings['sections']:
        tbl = settings['sections'][section_key]['table']
        conn.execute(f"UPDATE {tbl} SET accounted='محاسب' WHERE buyer=? AND accounted!='محاسب'", (name,))
    else:
        for cfg in settings['sections'].values():
            tbl = cfg['table']
            conn.execute(f"UPDATE {tbl} SET accounted='محاسب' WHERE buyer=? AND accounted!='محاسب'", (name,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/buyers/<name>/invoice')
def api_buyer_invoice(name):
    settings = load_settings()
    section_key = request.args.get('section', '')
    date_from = request.args.get('from', '')
    date_to = request.args.get('to', '')
    conn = get_db()
    all_rows = []
    for key, cfg in settings['sections'].items():
        if section_key and key != section_key:
            continue
        tbl = cfg['table']
        q = f"SELECT *, '{cfg['display']}' as section_name FROM {tbl} WHERE buyer=?"
        params = [name]
        if date_from:
            q += " AND date>=?"; params.append(date_from)
        if date_to:
            q += " AND date<=?"; params.append(date_to)
        q += " ORDER BY date"
        rows = [dict(r) for r in conn.execute(q, params).fetchall()]
        all_rows.extend(rows)

    # بيانات الدفعات
    payments = [dict(p) for p in conn.execute(
        "SELECT * FROM payments WHERE buyer=? ORDER BY date", (name,)).fetchall()]

    buyer_info = conn.execute("SELECT * FROM buyers WHERE name=?", (name,)).fetchone()
    conn.close()

    total = sum(r['total'] or 0 for r in all_rows)
    paid_amt = sum(r['total'] or 0 for r in all_rows if r['accounted'] == 'محاسب')
    paid_payments = sum(p['amount'] or 0 for p in payments)
    unpaid = total - paid_amt

    html = _generate_invoice_html(name, all_rows, total, paid_amt, unpaid,
                                   date_from, date_to, settings, payments)
    return html

def _generate_invoice_html(buyer, rows, total, paid, unpaid, fd, td, settings, payments):
    farm = settings.get('farm_name', 'مزرعتي')
    rows_html = ''
    for i, r in enumerate(rows, 1):
        status_cls = 'paid-r' if r['accounted'] == 'محاسب' else 'unpaid-r'
        badge_cls = 'badge-paid' if r['accounted'] == 'محاسب' else 'badge-unpaid'
        price_str = f"{r['price']:.2f}" if r.get('price') else '—'
        rows_html += f"""<tr class="{status_cls}">
<td>{i}</td><td>{r['date']}</td><td>{r.get('day','')}</td>
<td><strong>{r['product']}</strong></td>
<td>{r['qty']}</td><td>{r.get('unit','')}</td>
<td>{price_str}</td><td><strong>{r['total']:,.2f} ₪</strong></td>
<td><span class="badge {badge_cls}">{r['accounted']}</span></td>
</tr>"""
    fd_str = fd or (rows[0]['date'] if rows else '—')
    td_str = td or (rows[-1]['date'] if rows else '—')
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    return f"""<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>فاتورة {buyer}</title>
<style>
* {{margin:0;padding:0;box-sizing:border-box}}
body {{font-family:Tahoma,Arial,sans-serif;background:#f5f5f5;direction:rtl}}
.invoice {{max-width:900px;margin:20px auto;background:#fff;box-shadow:0 4px 20px rgba(0,0,0,.1);border-radius:12px;overflow:hidden}}
.inv-header {{background:linear-gradient(135deg,#1B4332,#2D6A4F);color:#fff;padding:30px;text-align:center}}
.inv-header h1 {{font-size:28px;margin-bottom:8px}}
.inv-header .sub {{color:#95D5B2;font-size:14px}}
.inv-meta {{display:grid;grid-template-columns:1fr 1fr;background:#f8fdf8}}
.meta-block {{padding:18px 24px;border-bottom:3px solid #2ECC71}}
.meta-label {{color:#52796F;font-size:11px;font-weight:600;margin-bottom:4px}}
.meta-value {{font-size:17px;font-weight:700;color:#0A1A10}}
table {{width:100%;border-collapse:collapse;margin:0}}
thead {{background:linear-gradient(135deg,#0A1A10,#1B4332)}}
th {{padding:11px;color:#74C69D;font-size:12px;font-weight:700;text-align:center}}
td {{padding:9px 11px;border-bottom:1px solid #eef5ee;font-size:13px;text-align:center}}
.paid-r {{background:#f0fdf4}}.unpaid-r {{background:#fff5f5}}
.badge {{padding:3px 10px;border-radius:20px;font-size:11px;font-weight:700}}
.badge-paid {{background:#d1fae5;color:#065f46}}
.badge-unpaid {{background:#fee2e2;color:#991b1b}}
.totals {{padding:24px;background:#f8fdf8;border-top:2px solid #e0efe0}}
.totals-grid {{display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px}}
.tot-card {{background:#fff;border-radius:10px;padding:14px;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,.06)}}
.tot-card .lbl {{color:#52796F;font-size:11px;margin-bottom:6px}}
.tot-card .val {{font-size:22px;font-weight:900}}
.green {{color:#1B5E20}}.red {{color:#C62828}}.dark {{color:#0A1A10}}
.footer {{background:#0A1A10;color:#4A7C59;text-align:center;padding:14px;font-size:12px}}
.footer strong {{color:#2ECC71}}
.print-btn {{background:#2ECC71;color:#fff;border:none;padding:10px 24px;border-radius:8px;cursor:pointer;font-size:14px;margin:16px;float:left}}
@media print {{.print-btn{{display:none}}.invoice{{box-shadow:none;margin:0}}}}
@media(max-width:600px) {{.totals-grid{{grid-template-columns:1fr}}.inv-meta{{grid-template-columns:1fr}}}}
</style>
</head>
<body>
<div class="invoice">
<div class="inv-header">
  <h1>🧾 فاتورة مشتري</h1>
  <div class="sub">{farm} — نظام محصولي 🌿</div>
</div>
<div class="inv-meta">
  <div class="meta-block">
    <div class="meta-label">المشتري</div>
    <div class="meta-value">👤 {buyer}</div>
  </div>
  <div class="meta-block">
    <div class="meta-label">الفترة</div>
    <div class="meta-value">📅 {fd_str} — {td_str}</div>
  </div>
  <div class="meta-block">
    <div class="meta-label">عدد السجلات</div>
    <div class="meta-value">{len(rows)} سجل</div>
  </div>
  <div class="meta-block">
    <div class="meta-label">تاريخ الإصدار</div>
    <div class="meta-value" style="font-size:14px">{now}</div>
  </div>
</div>
<button class="print-btn" onclick="window.print()">🖨️ طباعة</button>
<table>
<thead><tr>
<th>#</th><th>التاريخ</th><th>اليوم</th><th>المحصول</th>
<th>الكمية</th><th>الوحدة</th><th>السعر</th><th>المجموع</th><th>الحالة</th>
</tr></thead>
<tbody>{rows_html}</tbody>
</table>
<div class="totals">
<div class="totals-grid">
  <div class="tot-card"><div class="lbl">إجمالي الفاتورة</div><div class="val dark">{total:,.2f} ₪</div></div>
  <div class="tot-card"><div class="lbl">المبلغ المحصّل</div><div class="val green">{paid:,.2f} ₪</div></div>
  <div class="tot-card"><div class="lbl">المبلغ المتبقي</div><div class="val {'red' if unpaid>0 else 'green'}">{unpaid:,.2f} ₪</div></div>
</div>
</div>
<div class="footer"><strong>🌿 نظام محصولي</strong> — فاتورة {buyer} — {datetime.now().strftime('%Y-%m-%d')}</div>
</div>
</body></html>"""

@app.route('/api/payments', methods=['GET'])
def api_get_payments():
    buyer = request.args.get('buyer', '')
    conn = get_db()
    q = "SELECT * FROM payments WHERE 1=1"
    params = []
    if buyer:
        q += " AND buyer=?"; params.append(buyer)
    q += " ORDER BY date DESC"
    rows = [dict(r) for r in conn.execute(q, params).fetchall()]
    conn.close()
    return jsonify(rows)

@app.route('/api/payments', methods=['POST'])
def api_add_payment():
    data = request.json
    conn = get_db()
    pid = new_id()
    conn.execute("""INSERT INTO payments (id,buyer,amount,date,notes,section,created_at)
        VALUES (?,?,?,?,?,?,?)""",
        (pid, data['buyer'], safe_float(data.get('amount',0)),
         data.get('date', date.today().isoformat()),
         data.get('notes',''), data.get('section',''),
         datetime.now().isoformat()))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': pid})

@app.route('/api/expenses', methods=['GET'])
def api_get_expenses():
    conn = get_db()
    rows = [dict(r) for r in conn.execute("SELECT * FROM expenses ORDER BY date DESC").fetchall()]
    conn.close()
    return jsonify(rows)

@app.route('/api/expenses', methods=['POST'])
def api_add_expense():
    data = request.json
    conn = get_db()
    eid = new_id()
    conn.execute("""INSERT INTO expenses (id,date,category,amount,notes,section,created_at)
        VALUES (?,?,?,?,?,?,?)""",
        (eid, data.get('date', date.today().isoformat()),
         data.get('category',''), safe_float(data.get('amount',0)),
         data.get('notes',''), data.get('section',''),
         datetime.now().isoformat()))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': eid})

@app.route('/api/report')
def api_report():
    settings = load_settings()
    section_key = request.args.get('section', '')
    date_from = request.args.get('from', '')
    date_to = request.args.get('to', '')
    buyer = request.args.get('buyer', '')
    product = request.args.get('product', '')

    conn = get_db()
    all_rows = []
    for key, cfg in settings['sections'].items():
        if section_key and key != section_key:
            continue
        tbl = cfg['table']
        q = f"SELECT *, '{cfg['display']}' as section_name FROM {tbl} WHERE 1=1"
        params = []
        if date_from:
            q += " AND date>=?"; params.append(date_from)
        if date_to:
            q += " AND date<=?"; params.append(date_to)
        if buyer:
            q += " AND buyer=?"; params.append(buyer)
        if product:
            q += " AND product=?"; params.append(product)
        rows = [dict(r) for r in conn.execute(q, params).fetchall()]
        all_rows.extend(rows)

    all_rows.sort(key=lambda x: x.get('date',''), reverse=True)
    total = sum(r['total'] or 0 for r in all_rows)
    paid = sum(r['total'] or 0 for r in all_rows if r['accounted'] == 'محاسب')

    # إحصائيات
    by_product = defaultdict(lambda: {'qty': 0, 'total': 0})
    by_buyer = defaultdict(lambda: {'total': 0, 'unpaid': 0})
    by_date = defaultdict(float)
    for r in all_rows:
        if r.get('product'):
            by_product[r['product']]['qty'] += r['qty'] or 0
            by_product[r['product']]['total'] += r['total'] or 0
        if r.get('buyer'):
            by_buyer[r['buyer']]['total'] += r['total'] or 0
            if r['accounted'] != 'محاسب':
                by_buyer[r['buyer']]['unpaid'] += r['total'] or 0
        if r.get('date'):
            by_date[r['date']] += r['total'] or 0

    conn.close()
    return jsonify({
        'rows': all_rows[:500],
        'total': total, 'paid': paid, 'unpaid': total - paid,
        'count': len(all_rows),
        'by_product': [{'name': k, **v} for k, v in sorted(by_product.items(), key=lambda x: -x[1]['total'])],
        'by_buyer': [{'name': k, **v} for k, v in sorted(by_buyer.items(), key=lambda x: -x[1]['total'])[:10]],
        'by_date': [{'date': k, 'total': v} for k, v in sorted(by_date.items())[-30:]]
    })

@app.route('/api/export/excel')
def api_export_excel():
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return jsonify({'error': 'openpyxl غير متوفر'}), 500

    settings = load_settings()
    section_key = request.args.get('section', '')
    date_from = request.args.get('from', '')
    date_to = request.args.get('to', '')

    conn = get_db()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    green_fill = PatternFill("solid", fgColor="1B4332")
    header_font = Font(bold=True, color="FFFFFF", name='Arial')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for key, cfg in settings['sections'].items():
        if section_key and key != section_key:
            continue
        tbl = cfg['table']
        q = f"SELECT * FROM {tbl} WHERE 1=1"
        params = []
        if date_from:
            q += " AND date>=?"; params.append(date_from)
        if date_to:
            q += " AND date<=?"; params.append(date_to)
        q += " ORDER BY date DESC"
        rows = [dict(r) for r in conn.execute(q, params).fetchall()]

        ws = wb.create_sheet(title=cfg['display'])
        ws.sheet_view.rightToLeft = True
        headers = ['التاريخ','اليوم','المحصول','الكمية','الوحدة','السعر','المجموع','المشتري','المحاسبة','ملاحظات']
        for i, h in enumerate(headers, 1):
            cell = ws.cell(1, i, h)
            cell.fill = green_fill
            cell.font = header_font
            cell.alignment = center
        ws.row_dimensions[1].height = 25

        for row in rows:
            ws.append([row.get('date',''), row.get('day',''), row.get('product',''),
                       row.get('qty',0), row.get('unit',''), row.get('price') or '',
                       row.get('total',0), row.get('buyer',''), row.get('accounted',''),
                       row.get('notes','')])

        # عمود المجاميع
        total = sum(r['total'] or 0 for r in rows)
        paid = sum(r['total'] or 0 for r in rows if r['accounted'] == 'محاسب')
        last = ws.max_row + 2
        ws.cell(last, 1, 'الإجمالي').font = Font(bold=True, name='Arial')
        ws.cell(last, 7, total).font = Font(bold=True, name='Arial', color='1B4332')
        ws.cell(last+1, 1, 'المحصّل').font = Font(bold=True, name='Arial')
        ws.cell(last+1, 7, paid).font = Font(bold=True, name='Arial', color='2D6A4F')
        ws.cell(last+2, 1, 'المتبقي').font = Font(bold=True, name='Arial')
        ws.cell(last+2, 7, total-paid).font = Font(bold=True, name='Arial', color='C62828')

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 14

    conn.close()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"محصولي_{date.today().isoformat()}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/backup', methods=['POST'])
def api_backup():
    try:
        stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        dest = os.path.join(BACKUP_DIR, f"backup_{stamp}.db")
        shutil.copy2(DB_FILE, dest)
        # احتفظ بآخر 10 نسخ
        backups = sorted([f for f in os.listdir(BACKUP_DIR) if f.endswith('.db')])
        for old in backups[:-10]:
            os.remove(os.path.join(BACKUP_DIR, old))
        return jsonify({'success': True, 'file': f"backup_{stamp}.db"})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/settings', methods=['GET'])
def api_get_settings():
    return jsonify(load_settings())

@app.route('/api/settings', methods=['POST'])
def api_save_settings():
    data = request.json
    s = load_settings()
    if 'farm_name' in data:
        s['farm_name'] = data['farm_name']
    if 'products' in data:
        s['products'] = data['products']
    if 'worker_ratio' in data:
        s['worker_ratio'] = float(data['worker_ratio'])
    if 'sections' in data:
        for key, sec in data['sections'].items():
            if key in s['sections']:
                s['sections'][key].update(sec)
    save_settings(s)
    init_db()
    return jsonify({'success': True})

@app.route('/api/whatsapp/<buyer_name>')
def api_whatsapp(buyer_name):
    settings = load_settings()
    import urllib.parse
    conn = get_db()
    b = conn.execute("SELECT phone FROM buyers WHERE name=?", (buyer_name,)).fetchone()
    debt = 0
    for cfg in settings['sections'].values():
        tbl = cfg['table']
        r = conn.execute(f"SELECT SUM(total) FROM {tbl} WHERE buyer=? AND accounted!='محاسب'", (buyer_name,)).fetchone()
        debt += r[0] or 0
    conn.close()
    phone = (b['phone'] if b else '') or ''
    phone = re.sub(r'[^0-9]', '', phone)
    if phone.startswith('0'):
        phone = '970' + phone[1:]
    msg = f"السلام عليكم {buyer_name}\nرصيدك المتبقي: {debt:,.0f} شيقل\nنرجو التسوية في أقرب وقت\nشكراً — {settings.get('farm_name','مزرعتي')}"
    encoded = urllib.parse.quote(msg, safe='')
    url = f"https://wa.me/{phone}?text={encoded}" if phone else f"https://wa.me/?text={encoded}"
    return jsonify({'url': url, 'debt': debt, 'phone': phone})

@app.route('/api/seasons', methods=['GET'])
def api_get_seasons():
    section = request.args.get('section', '')
    conn = get_db()
    q = "SELECT * FROM seasons WHERE 1=1"
    params = []
    if section:
        q += " AND (section_key=? OR section_key='الكل')"
        params.append(section)
    q += " ORDER BY start_date DESC"
    rows = [dict(r) for r in conn.execute(q, params).fetchall()]
    conn.close()
    return jsonify(rows)

@app.route('/api/db_stats')
def api_db_stats():
    settings = load_settings()
    conn = get_db()
    stats = {}
    total_records = 0
    for key, cfg in settings['sections'].items():
        tbl = cfg['table']
        cnt = conn.execute(f"SELECT COUNT(*) FROM {tbl}").fetchone()[0]
        stats[key] = cnt
        total_records += cnt
    conn.close()
    db_size = os.path.getsize(DB_FILE) if os.path.exists(DB_FILE) else 0
    return jsonify({
        'sections': stats,
        'total': total_records,
        'db_size_kb': round(db_size / 1024, 1),
        'backups': len([f for f in os.listdir(BACKUP_DIR) if f.endswith('.db')])
    })

if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
